"""
Gerador de planilhas Excel e CSV para relatórios do simulador atuarial
"""
import io
import csv
import uuid
import time
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from .models.report_models import ReportRequest, ReportResponse, ReportConfig
from ...utils.formatters import format_currency_br
from .abstract_report_generator import AbstractReportGenerator


class ExcelGenerator(AbstractReportGenerator):
    """Gerador de planilhas Excel e CSV com dados estruturados"""

    def __init__(self, config: ReportConfig = None, cache_dir: Path = None):
        super().__init__(config, cache_dir)

    @property
    def supported_formats(self) -> list[str]:
        """Formatos suportados pelo gerador Excel."""
        return ['xlsx', 'csv']

    @property
    def default_content_type(self) -> str:
        """Tipo MIME padrão para Excel."""
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def generate_excel(self, request: ReportRequest) -> ReportResponse:
        """
        Método público para gerar Excel (mantém compatibilidade).
        Delega para o método base generate_report.
        """
        return self.generate_report(request)

    def _generate_specific_report(self, request: ReportRequest, report_id: str) -> ReportResponse:
        """
        Implementação específica da geração Excel.
        """
        # Criar workbook
        wb = openpyxl.Workbook()

        # Remover aba padrão
        wb.remove(wb.active)

        # Criar abas com dados
        self._create_input_data_sheet(wb, request)
        self._create_results_sheet(wb, request)
        self._create_projections_sheet(wb, request)
        self._create_assumptions_sheet(wb, request)

        # Salvar arquivo
        file_path = self.cache_dir / f"dados_simulacao_{report_id}.xlsx"
        wb.save(file_path)

        # Obter tamanho do arquivo
        file_size = self._get_file_size(file_path)

        return self._create_success_response(
            report_id=report_id,
            generation_time_ms=0,  # Será calculado pela classe base
            file_path=str(file_path),
            file_size=file_size,
            content_type=self.default_content_type,
            message="Planilha Excel gerada com sucesso"
        )

    def generate_csv(self, request: ReportRequest) -> ReportResponse:
        """
        Gerar arquivo CSV com dados principais
        """
        start_time = time.time()
        report_id = str(uuid.uuid4())

        try:
            # Criar dados para CSV
            csv_data = self._prepare_csv_data(request)

            # Salvar arquivo CSV
            file_path = self.cache_dir / f"dados_simulacao_{report_id}.csv"

            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')

                # Escrever dados
                for row in csv_data:
                    writer.writerow(row)

            generation_time = int((time.time() - start_time) * 1000)
            file_size = file_path.stat().st_size

            return ReportResponse(
                success=True,
                message="Arquivo CSV gerado com sucesso",
                file_path=str(file_path),
                file_size=file_size,
                generation_time_ms=generation_time,
                report_id=report_id,
                content_type="text/csv"
            )

        except Exception as e:
            return ReportResponse(
                success=False,
                message=f"Erro ao gerar CSV: {str(e)}",
                generation_time_ms=int((time.time() - start_time) * 1000),
                report_id=report_id
            )

    def _create_input_data_sheet(self, wb: openpyxl.Workbook, request: ReportRequest):
        """Criar aba com dados de entrada"""
        ws = wb.create_sheet("Dados de Entrada")

        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        # Cabeçalho
        ws.append(["Parâmetro", "Valor", "Descrição"])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Dados do participante
        ws.append(["DADOS DO PARTICIPANTE", "", ""])
        ws.append(["Idade Atual", request.state.age, "Anos"])
        ws.append(["Gênero", "Masculino" if request.state.gender == "M" else "Feminino", ""])
        ws.append(["Salário Mensal", request.state.salary, "R$"])
        ws.append(["Idade de Aposentadoria", request.state.retirement_age, "Anos"])

        # Parâmetros do plano
        ws.append(["", "", ""])
        ws.append(["PARÂMETROS DO PLANO", "", ""])
        ws.append(["Tipo de Plano", "BD" if request.state.plan_type == "BD" else "CD", "Benefício Definido / Contribuição Definida"])
        ws.append(["Saldo Inicial", request.state.initial_balance, "R$"])
        ws.append(["Taxa de Contribuição", request.state.contribution_rate, "%"])

        if request.state.plan_type == "BD":
            ws.append(["Taxa de Benefício", request.state.accrual_rate, "% do salário por ano"])
            if hasattr(request.state, 'target_benefit') and request.state.target_benefit:
                ws.append(["Benefício Alvo", request.state.target_benefit, "R$"])
        else:
            if hasattr(request.state, 'accumulation_rate') and request.state.accumulation_rate:
                ws.append(["Taxa de Acumulação", request.state.accumulation_rate * 100, "% a.a."])
            if hasattr(request.state, 'conversion_rate') and request.state.conversion_rate:
                ws.append(["Taxa de Conversão", request.state.conversion_rate * 100, "% a.a."])

        # Base atuarial
        ws.append(["", "", ""])
        ws.append(["BASE ATUARIAL", "", ""])
        ws.append(["Tábua de Mortalidade", request.state.mortality_table, ""])
        ws.append(["Suavização de Mortalidade", request.state.mortality_aggravation, "%"])
        ws.append(["Taxa de Desconto", request.state.discount_rate * 100, "% a.a."])
        ws.append(["Crescimento Salarial Real", request.state.salary_growth_real * 100, "% a.a."])

        # Configurações
        ws.append(["", "", ""])
        ws.append(["CONFIGURAÇÕES", "", ""])
        ws.append(["Regime de Pagamento", request.state.payment_timing.title(), ""])
        ws.append(["Salários por Ano", request.state.salary_months_per_year, "Meses"])
        ws.append(["Benefícios por Ano", request.state.benefit_months_per_year, "Meses"])
        ws.append(["Taxa de Administração", request.state.admin_fee_rate * 100, "% a.a."])

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40

        # Aplicar formatação monetária nas células com valores em R$
        last_row = ws.max_row
        self._apply_currency_format(ws, f"B4:B4")  # Salário Mensal
        self._apply_currency_format(ws, f"B10:B10")  # Saldo Inicial

        # Benefício Alvo - buscar dinamicamente
        for i in range(2, last_row + 1):
            if ws[f'A{i}'].value and "Benefício Alvo" in str(ws[f'A{i}'].value):
                self._apply_currency_format(ws, f"B{i}:B{i}")

        self._apply_borders(ws)

    def _create_results_sheet(self, wb: openpyxl.Workbook, request: ReportRequest):
        """Criar aba com resultados principais"""
        ws = wb.create_sheet("Resultados Principais")

        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Cabeçalho
        ws.append(["Indicador", "Valor", "Unidade"])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Resultados baseados no tipo de plano
        if request.state.plan_type == "BD":
            ws.append(["RMBA (Reserva Matemática)", request.results.rmba or 0, "R$"])
            ws.append(["RMBC (Custo do Benefício)", request.results.rmbc or 0, "R$"])
            ws.append(["Déficit/Superávit", request.results.deficit_surplus or 0, "R$"])
            ws.append(["Taxa de Contribuição Necessária", request.results.required_contribution_rate or 0, "%"])
        else:
            ws.append(["Saldo Final Acumulado", request.results.individual_balance or 0, "R$"])
            ws.append(["Renda Mensal Projetada", request.results.monthly_income_cd or 0, "R$"])

        # Indicadores comuns
        ws.append(["Taxa de Reposição", (request.results.replacement_ratio or 0) / 100, "%"])

        # Dados da simulação
        ws.append(["", "", ""])
        ws.append(["INFORMAÇÕES DA SIMULAÇÃO", "", ""])
        ws.append(["Data de Geração", datetime.now().strftime("%d/%m/%Y %H:%M"), ""])
        ws.append(["Versão do Simulador", "1.0", ""])

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

        # Aplicar formatação monetária e percentual
        if request.state.plan_type == "BD":
            self._apply_currency_format(ws, "B2:B4")  # RMBA, RMBC, Déficit/Superávit
            self._apply_percentage_format(ws, "B5:B5")  # Taxa de Contribuição Necessária
        else:
            self._apply_currency_format(ws, "B2:B3")  # Saldo Final, Renda Mensal

        # Taxa de Reposição (sempre percentual)
        last_row = ws.max_row
        for i in range(2, last_row + 1):
            if ws[f'A{i}'].value and "Taxa de Reposição" in str(ws[f'A{i}'].value):
                ws[f'B{i}'].number_format = '0.00%'

        self._apply_borders(ws)

    def _create_projections_sheet(self, wb: openpyxl.Workbook, request: ReportRequest):
        """Criar aba com projeções anuais"""
        ws = wb.create_sheet("Projeções Anuais")

        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Cabeçalho baseado no tipo de plano
        if request.state.plan_type == "BD":
            headers = ["Ano", "Idade", "Salário Anual", "Contribuições", "RMBA", "Prob. Sobrevivência Cumulativa (%)"]
        else:
            headers = ["Ano", "Idade", "Salário Anual", "Contribuições", "Saldo Acumulado", "Prob. Sobrevivência Cumulativa (%)"]

        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Dados das projeções
        projection_years = getattr(request.results, 'projection_years', [])
        current_year = datetime.now().year

        if projection_years and len(projection_years) > 0:
            for i, year in enumerate(projection_years):
                row_data = [
                    current_year + i,  # Ano
                    request.state.age + i,  # Idade
                ]

                # Salário projetado
                if hasattr(request.results, 'projected_salaries') and request.results.projected_salaries:
                    salary = request.results.projected_salaries[i] if i < len(request.results.projected_salaries) else 0
                    row_data.append(salary)
                else:
                    row_data.append(0)

                # Contribuições
                if hasattr(request.results, 'projected_contributions') and request.results.projected_contributions:
                    contrib = request.results.projected_contributions[i] if i < len(request.results.projected_contributions) else 0
                    row_data.append(contrib)
                else:
                    row_data.append(0)

                # RMBA ou Saldo Acumulado
                if request.state.plan_type == "BD":
                    if hasattr(request.results, 'projected_rmba_evolution') and request.results.projected_rmba_evolution:
                        rmba = request.results.projected_rmba_evolution[i] if i < len(request.results.projected_rmba_evolution) else 0
                        row_data.append(rmba)
                    else:
                        row_data.append(0)
                else:
                    if hasattr(request.results, 'projected_balances') and request.results.projected_balances:
                        balance = request.results.projected_balances[i] if i < len(request.results.projected_balances) else 0
                        row_data.append(balance)
                    else:
                        row_data.append(0)

                # Probabilidade de sobrevivência
                if hasattr(request.results, 'survival_probabilities') and request.results.survival_probabilities:
                    prob = request.results.survival_probabilities[i] if i < len(request.results.survival_probabilities) else 0
                    row_data.append(prob)  # Valor decimal (formatação % aplicada depois)
                else:
                    row_data.append(0)

                ws.append(row_data)

        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Aplicar formatação monetária e percentual nas projeções
        if projection_years and len(projection_years) > 0:
            last_row = ws.max_row
            # Colunas monetárias: Salário Anual (C), Contribuições (D), RMBA/Saldo (E)
            self._apply_currency_format(ws, f"C2:E{last_row}")
            # Coluna de probabilidade: Probabilidade Sobrevivência (F)
            self._apply_percentage_format(ws, f"F2:F{last_row}")

        self._apply_borders(ws)

    def _create_assumptions_sheet(self, wb: openpyxl.Workbook, request: ReportRequest):
        """Criar aba com premissas técnicas"""
        ws = wb.create_sheet("Premissas Técnicas")

        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Cabeçalho
        ws.append(["Premissa", "Valor", "Observações"])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Premissas demográficas
        ws.append(["PREMISSAS DEMOGRÁFICAS", "", ""])
        ws.append(["Tábua de Mortalidade", request.state.mortality_table, "Base técnica para cálculo de sobrevivência"])

        aggravation_text = "Sem suavização" if request.state.mortality_aggravation == 0 else f"Suavização de {request.state.mortality_aggravation}%"
        ws.append(["Suavização de Mortalidade", aggravation_text, "Ajuste aplicado à tábua base"])

        # Premissas econômicas
        ws.append(["", "", ""])
        ws.append(["PREMISSAS ECONÔMICAS", "", ""])
        ws.append(["Taxa de Desconto", f"{request.state.discount_rate * 100:.2f}% a.a.", "Taxa real, livre de inflação"])
        ws.append(["Crescimento Salarial Real", f"{request.state.salary_growth_real * 100:.2f}% a.a.", "Crescimento acima da inflação"])
        ws.append(["Taxa de Administração", f"{request.state.admin_fee_rate * 100:.2f}% a.a.", "Custos administrativos anuais"])

        # Premissas do plano
        ws.append(["", "", ""])
        ws.append(["PREMISSAS DO PLANO", "", ""])
        ws.append(["Regime de Pagamento", request.state.payment_timing.title(), "Momento do pagamento das contribuições"])
        ws.append(["Indexação de Benefícios", request.state.benefit_indexation or "Não especificada", ""])
        ws.append(["Indexação de Contribuições", request.state.contribution_indexation or "Não especificada", ""])

        # Metodologia
        ws.append(["", "", ""])
        ws.append(["METODOLOGIA", "", ""])
        calc_method = getattr(request.state, 'calculation_method', 'PUC')
        ws.append(["Método de Cálculo", calc_method, "Metodologia atuarial utilizada"])
        ws.append(["Projeção", f"{request.state.projection_years} anos", "Período de projeção da simulação"])

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50

        self._apply_borders(ws)

    def _apply_borders(self, ws):
        """Aplicar bordas às células com dados"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

    def _apply_currency_format(self, ws, cell_range: str):
        """Aplicar formatação monetária brasileira"""
        for row in ws[cell_range]:
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value != 0:
                    cell.number_format = 'R$ #,##0.00'

    def _apply_percentage_format(self, ws, cell_range: str):
        """Aplicar formatação de percentual"""
        for row in ws[cell_range]:
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00%'

    def _prepare_csv_data(self, request: ReportRequest) -> List[List[str]]:
        """Preparar dados para arquivo CSV"""
        csv_data = []

        # Cabeçalho
        csv_data.append(["Simulação Atuarial Individual - Dados Exportados"])
        csv_data.append([f"Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        csv_data.append([])

        # Dados de entrada
        csv_data.append(["DADOS DE ENTRADA"])
        csv_data.append(["Parâmetro", "Valor"])
        csv_data.append(["Idade Atual", str(request.state.age)])
        csv_data.append(["Gênero", "Masculino" if request.state.gender == "M" else "Feminino"])
        csv_data.append(["Salário Mensal", format_currency_br(request.state.salary)])
        csv_data.append(["Tipo de Plano", request.state.plan_type])
        csv_data.append(["Taxa de Contribuição", f"{request.state.contribution_rate}%"])
        csv_data.append([])

        # Resultados principais
        csv_data.append(["RESULTADOS PRINCIPAIS"])
        csv_data.append(["Indicador", "Valor"])

        if request.state.plan_type == "BD":
            csv_data.append(["RMBA", format_currency_br(request.results.rmba or 0)])
            csv_data.append(["RMBC", format_currency_br(request.results.rmbc or 0)])
            csv_data.append(["Déficit/Superávit", format_currency_br(request.results.deficit_surplus or 0)])
        else:
            csv_data.append(["Saldo Final", format_currency_br(request.results.individual_balance or 0)])
            csv_data.append(["Renda Mensal", format_currency_br(request.results.monthly_income_cd or 0)])

        csv_data.append(["Taxa de Reposição", f"{(request.results.replacement_ratio or 0):.3f}"])

        return csv_data
